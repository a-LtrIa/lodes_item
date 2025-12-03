import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

def extract_items_from_page():
    """
    从指定的 PRTS 页面中提取符合条件的 <div> 标签内容。
    
    参数:
        url (str): PRTS 页面的 URL。
    
    返回:
        list: 包含 (title, span_text) 元组的列表。
    """
    url = "https://prts.wiki/w/CHAR#?rarity=1-6"
    # 发送 HTTP 请求
    response = requests.get(url)
    if response.status_code != 200:
        print("Failed to retrieve the webpage")
        return []

    # 使用 BeautifulSoup 解析 HTML
    soup = BeautifulSoup(response.text, 'html.parser')

    # 查找所有符合条件的 <div> 标签
    items = soup.find_all('a', attrs={'data-v-bb362039': True})
    print(f"Found {len(items)} items")

    # 提取数据
    results = []
    for item in items:
        # 查找 <div> 标签
        div_tags = item.find_all('div', attrs={'data-v-bb362039': True})
        if div_tags:
            # 获取第一个符合条件的 <div> 标签的文本内容
            div_text = div_tags[0].get_text(strip=True)
            results.append(div_text)  # 保存结果
            

    
    return results


result=extract_items_from_page()
print(result)
output_file = "干员一览.xlsx"

# 写入 Excel 文件
workbook = load_workbook(output_file)
sheet = workbook.active

for i, item in enumerate(result):
    sheet.cell(row=i+2, column=1, value=item)

workbook.save(output_file)